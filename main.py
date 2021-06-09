from flask import Flask, render_template, request, redirect, url_for, flash
from flask_bootstrap import Bootstrap
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField, SelectField
from wtforms.fields.html5 import DateField
from wtforms.validators import DataRequired, Required
import os
from werkzeug.utils import secure_filename
import csv
import pandas as pd
from openpyxl import load_workbook
from datetime import date

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get("APP_CONFIG_SECRET_KEY")
app.config['UPLOAD_FOLDER'] = "./Archivos Excel"
Bootstrap(app)

# Create SQLite DATABASE:
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL",  "sqlite:///personas.db")
# Optional: But it will silence the deprecation warning in the console.
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


# Create Table
class Person(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(80), nullable=False)
    apellido = db.Column(db.String(80), nullable=False)
    nacionalidad = db.Column(db.String(50), nullable=False)
    fecha_contrato = db.Column(db.String(50), nullable=False)
    sexo = db.Column(db.String(50), nullable=False)


db.create_all()



@app.route("/")
def home():
    # READ ALL PEOPLE:
    all_people = Person.query.order_by(Person.nombre).all()
    # #This line loops through all the people
    # for i in range(len(all_people)):
    #     #This line gives each movie a new ranking reversed from their order in all_movies
    #     all_people[i].ranking = len(all_people) - i
    # # db.session.commit()
    return render_template("index.html", people=all_people)



# UPDATE RATING AND REVIEW
class AddPersonForm(FlaskForm):
    nombre = StringField("Nombre(s)", validators=[DataRequired()])
    apellido = StringField("Apellido(s)", validators=[DataRequired()])
    nacionalidad = StringField("Nacionalidad", validators=[DataRequired()], render_kw={"placeholder": "ej. Mexicana"})
    fecha_contrato = DateField('Fecha Contrato', validators=[Required()], format = '%Y-%m-%d')
    sexo = SelectField(u'Sexo', choices=[('Hombre', 'Hombre'), ('Mujer', 'Mujer')])
    submit = SubmitField("Add Person")


@app.route("/add", methods=["GET", "POST"])
def add_person():
    form = AddPersonForm()
    print(form.validate())
    if request.method == "POST" and form.validate() == False:
        flash('All Fields are required.')
        return redirect(url_for('add_person'))
    if form.validate_on_submit():
        data = form.data
        fecha = str(data["fecha_contrato"])
        new_person = Person(
            nombre = data["nombre"].title(),
            apellido = data["apellido"].title(),
            nacionalidad = data["nacionalidad"].title(),
            fecha_contrato = fecha.split("-")[1]+'/'+fecha.split("-")[2]+'/'+fecha.split("-")[0],
            sexo = data["sexo"]
        )
        db.session.add(new_person)
        db.session.commit()
        flash('Persona agregada exitosamente')
        return redirect(url_for("home"))
        print("--------------")
    return render_template("add.html", form=form)


@app.route("/edit", methods=["GET", "POST"])
def edit_person():
    form = AddPersonForm()
    person_id = request.args.get("id")
    person = Person.query.get(person_id)
    print(person.fecha_contrato)
    fecha_previa = person.fecha_contrato
    if person.fecha_contrato != None:
        fecha_previa = person.fecha_contrato.split("/")[2]+"-"+person.fecha_contrato.split("/")[0]+"-"+person.fecha_contrato.split("/")[1]
    print(fecha_previa)
    print("###############")
    if request.method == "POST":
        print("HEY")
        data = form.data
        fecha = str(data["fecha_contrato"])
        print(fecha)
        person.nombre = data["nombre"].title()
        person.apellido = data["apellido"].title()
        person.nacionalidad = data["nacionalidad"].title()
        person.fecha_contrato = fecha.split("-")[1]+'/'+fecha.split("-")[2]+'/'+fecha.split("-")[0]
        person.sexo = data["sexo"]
        db.session.commit()
        flash('Datos editados exitosamente')
        # return redirect(url_for('add_person'))
        return redirect(url_for("home"))
    return render_template("edit.html", person=person, form=form, fecha_previa=fecha_previa)


@app.route("/delete")
def delete_person():
    person_id = request.args.get("id")
    # DELETE PERSON
    person_to_delete = Person.query.get(person_id)
    db.session.delete(person_to_delete)
    db.session.commit()
    return redirect(url_for("home"))


@app.route("/uploader", methods=["POST"])
def uploader():
    if request.method == "POST":
        f = request.files['archivo']
        filename = secure_filename(f.filename)
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        # file_path= f"folder/.xlsx"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print(file_path)
        # dfs = pd.read_excel(file_path)
        # print(dfs.head(5))
        wb2 = load_workbook(file_path)
        print(wb2.sheetnames)
        ws = wb2.active
        # https://openpyxl.readthedocs.io/en/stable/tutorial.html
        # for row in ws.iter_rows(min_row=1, max_col=5, max_row=14, values_only=True):
        #     print(row)
        new_data = []
        count = 0
        agregados = False
        for row in ws.values:
            list_row = list(row)    # convert tuple to list
            count += 1
            if list_row[0] == None:
                break
            if count != 1:
                if list_row[3] != None:
                    date_str = (str(list_row[3]).split(" ")) # to get as ['2018-11-15', '00:00:00']
                    fecha = date_str[0].split("-")
                    list_row[3] = fecha[1] + "/" + fecha[2] + "/" + fecha[0] # rename the row 3
                new_data.append(list_row)
            new_person = Person(
            nombre = list_row[0].title(),
            apellido = list_row[1].title(),
            nacionalidad = list_row[2].title(),
            fecha_contrato = list_row[3],
            sexo = list_row[4]
            )
            db.session.add(new_person)
            db.session.commit()
            agregados = True
        if agregados:
            flash('Datos agregados exitosamente')
        return redirect(url_for("home"))


if __name__ == '__main__':
    app.run(debug=True)